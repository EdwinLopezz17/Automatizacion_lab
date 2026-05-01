import io
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
EMPTY_FILL = PatternFill(fill_type=None)
CORRECT_FILL = PatternFill("solid", fgColor="C6EFCE")
WRONG_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BORDER_THIN = Border(
    left=Side(style="thin",   color="B8CCE4"),
    right=Side(style="thin",  color="B8CCE4"),
    top=Side(style="thin",    color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)
SUMMARY_TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
SUMMARY_TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)

DATE_COLS_CESADOS = {
    "Fecha de Cese",
    "Usr Exactus Ultimo Login",
    "Usr SDP Ultimo Login",
    "DB SIT Ultimo Login",
    "DB SDP Ultimo Login",
    "DB Exactus Ultimo Login",
    "Entra ID Ultimo Login",
    "Ultimo Login AD Nipa",
}

def _style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def _style_data_rows(ws, start_row: int, end_row: int, n_cols: int,
                     validation_cols: Optional[dict] = None):
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.fill = EMPTY_FILL

def _auto_col_width(ws, min_w: int = 10, max_w: int = 45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value or "")) for c in col if c.value is not None]
        width = min(max(max(lengths, default=0) + 4, min_w), max_w)
        ws.column_dimensions[col_letter].width = width

def _df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame,
                 validation_cols: Optional[dict] = None,
                 date_cols: Optional[set] = None,
                 date_fmt: str = "DD/MM/YYYY"):

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 30

    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=col)
    _style_header_row(ws, 1, len(df.columns))

    date_idx = (
        {ci for ci, col in enumerate(df.columns, 1) if col in date_cols}
        if date_cols else set()
    )

    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            if ci in date_idx and val not in (None, ""):
                if hasattr(val, "year"):
                    cell_val = val
                else:
                    parsed   = pd.to_datetime(val, errors="coerce", dayfirst=True)
                    cell_val = parsed.to_pydatetime() if not pd.isnull(parsed) else None
                if cell_val is not None and hasattr(cell_val, "tzinfo") and cell_val.tzinfo is not None:
                    cell_val = cell_val.replace(tzinfo=None)
                cell = ws.cell(row=ri, column=ci, value=cell_val)
                if cell_val is not None:
                    cell.number_format = date_fmt
            else:
                cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
                if val and hasattr(val, "date") and not isinstance(val, str):
                    cell.number_format = "yyyy-mm-dd"

    if len(df) > 0:
        _style_data_rows(ws, 2, len(df) + 1, len(df.columns), validation_cols)

    _auto_col_width(ws)
    return ws


def _crear_wb_vacio() -> Workbook:
    wb = Workbook()
    for nombre in list(wb.sheetnames):
        del wb[nombre]
    return wb

def wb_to_buffer(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
